from rest_framework import generics, status, permissions, filters
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from django_filters.rest_framework import DjangoFilterBackend
from django.db.models import Sum, Q, F, DecimalField, Value
from django.db.models.functions import Coalesce
from django.utils import timezone
from datetime import datetime, date, timedelta
from .models import Payment, Expense, PaymentHistory
from .serializers import (PaymentSerializer, PaymentListSerializer, PaymentHistorySerializer,
                         ExpenseSerializer, ExpenseListSerializer, FinancialSummarySerializer)
from django.http import HttpResponse
import io
from openpyxl import Workbook
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4


class PaymentListCreateView(generics.ListCreateAPIView):
    """List and create payments"""
    queryset = Payment.objects.all()
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['status', 'payment_method', 'branch', 'created_at', 'payment_date']
    search_fields = ['customer__first_name', 'customer__last_name', 'customer__phone']
    ordering_fields = ['amount', 'created_at', 'payment_date']
    ordering = ['-created_at']
    
    def get_serializer_class(self):
        if self.request.method == 'GET':
            return PaymentListSerializer
        return PaymentSerializer


class PaymentDetailView(generics.RetrieveUpdateDestroyAPIView):
    """Retrieve, update or delete payment"""
    queryset = Payment.objects.all()
    serializer_class = PaymentSerializer
    permission_classes = [permissions.IsAuthenticated]


class ExpenseListCreateView(generics.ListCreateAPIView):
    """List and create expenses"""
    queryset = Expense.objects.all()
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['category', 'branch', 'expense_date']
    search_fields = ['title', 'description']
    ordering_fields = ['amount', 'expense_date', 'created_at']
    ordering = ['-expense_date']
    
    def get_serializer_class(self):
        if self.request.method == 'GET':
            return ExpenseListSerializer
        return ExpenseSerializer


class ExpenseDetailView(generics.RetrieveUpdateDestroyAPIView):
    """Retrieve, update or delete expense"""
    queryset = Expense.objects.all()
    serializer_class = ExpenseSerializer
    permission_classes = [permissions.IsAuthenticated]


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def add_payment(request, pk):
    """Add payment to existing payment record"""
    try:
        payment = Payment.objects.get(pk=pk)
        amount = request.data.get('amount')
        payment_method = request.data.get('payment_method')
        notes = request.data.get('notes', '')
        
        if not amount or amount <= 0:
            return Response(
                {'error': 'Invalid amount'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Create payment history
        PaymentHistory.objects.create(
            payment=payment,
            amount=amount,
            payment_method=payment_method,
            notes=notes,
            created_by=request.user
        )
        
        # Update payment
        payment.paid_amount += amount
        if payment.paid_amount >= payment.amount:
            payment.status = 'paid'
        else:
            payment.status = 'partial'
        payment.save()
        
        serializer = PaymentSerializer(payment)
        return Response(serializer.data)
        
    except Payment.DoesNotExist:
        return Response(
            {'error': 'Payment not found'}, 
            status=status.HTTP_404_NOT_FOUND
        )


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def payment_history(request, pk):
    """Get payment history"""
    try:
        payment = Payment.objects.get(pk=pk)
        history = payment.payment_history.all()
        serializer = PaymentHistorySerializer(history, many=True)
        return Response(serializer.data)
    except Payment.DoesNotExist:
        return Response(
            {'error': 'Payment not found'}, 
            status=status.HTTP_404_NOT_FOUND
        )


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def financial_summary(request):
    """Get financial summary"""
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    filter_all = False
    if not start_date or not end_date:
        # If no range provided, show ALL-TIME summary to avoid confusion
        filter_all = True
        today = timezone.now().date()
        start_date = today.replace(day=1)  # placeholder to satisfy types below
        end_date = today
    
    # Convert string dates to date objects
    if isinstance(start_date, str):
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    if isinstance(end_date, str):
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    # Calculate revenue
    # Treat revenue as cash-based: sum of paid_amount (by payment_date if available, fallback to created_at)
    payments_qs = Payment.objects.all()
    if not filter_all:
        payments_qs = payments_qs.filter(
            Q(payment_date__range=[start_date, end_date]) |
            (Q(payment_date__isnull=True) & Q(created_at__date__range=[start_date, end_date]))
        )
    paid_payments = payments_qs.aggregate(
        total=Coalesce(Sum('paid_amount'), Value(0, output_field=DecimalField()))
    )['total'] or 0
    # Total quoted amount (orders) in period for context
    total_revenue = payments_qs.aggregate(
        total=Coalesce(Sum('amount'), Value(0, output_field=DecimalField()))
    )['total'] or 0
    # Pending equals sum of remaining_amount for pending/partial in period
    pending_base = payments_qs.filter(status__in=['pending', 'partial'])
    pending_payments = pending_base.aggregate(
        total=Coalesce(
            Sum(F('amount') - F('paid_amount'), output_field=DecimalField()),
            Value(0, output_field=DecimalField())
        )
    )['total'] or 0
    
    # Calculate expenses
    expenses_qs = Expense.objects.all()
    if not filter_all:
        expenses_qs = expenses_qs.filter(expense_date__range=[start_date, end_date])
    total_expenses = expenses_qs.aggregate(total=Coalesce(Sum('amount'), Value(0, output_field=DecimalField())))['total'] or 0
    
    summary = {
        'total_revenue': total_revenue,
        'total_expenses': total_expenses,
        'pending_payments': pending_payments,
        'paid_payments': paid_payments,
        'period_start': start_date,
        'period_end': end_date
    }
    
    serializer = FinancialSummarySerializer(summary)
    return Response(serializer.data)


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def revenue_by_service(request):
    """Get revenue breakdown by service"""
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    if not start_date or not end_date:
        today = timezone.now().date()
        start_date = today.replace(day=1)
        end_date = today
    
    if isinstance(start_date, str):
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    if isinstance(end_date, str):
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    from customers.models import Service
    
    services = Service.objects.all()
    revenue_data = []
    
    for service in services:
        total_amount = Payment.objects.filter(
            service=service,
            created_at__date__range=[start_date, end_date]
        ).aggregate(total=Sum('amount'))['total'] or 0
        
        paid_amount = Payment.objects.filter(
            service=service,
            created_at__date__range=[start_date, end_date],
            status='paid'
        ).aggregate(total=Sum('paid_amount'))['total'] or 0
        
        revenue_data.append({
            'service_name': service.name,
            'total_amount': total_amount,
            'paid_amount': paid_amount,
            'pending_amount': total_amount - paid_amount
        })
    
    return Response(revenue_data)


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def financial_stats(request):
    """Get financial statistics"""
    today = timezone.now().date()
    this_month = today.replace(day=1)
    
    stats = {
        # Total quoted amounts and expenses overall
        'total_revenue': Payment.objects.aggregate(
            total=Coalesce(Sum('amount'), Value(0, output_field=DecimalField()))
        )['total'] or 0,
        'total_expenses': Expense.objects.aggregate(
            total=Coalesce(Sum('amount'), Value(0, output_field=DecimalField()))
        )['total'] or 0,
        # Cash-based revenue this month (prefer payment_date)
        'this_month_revenue': Payment.objects.filter(
            Q(payment_date__gte=this_month) | (Q(payment_date__isnull=True) & Q(created_at__date__gte=this_month))
        ).aggregate(total=Coalesce(Sum('paid_amount'), Value(0, output_field=DecimalField())))['total'] or 0,
        'this_month_expenses': Expense.objects.filter(
            expense_date__gte=this_month
        ).aggregate(total=Coalesce(Sum('amount'), Value(0, output_field=DecimalField())))['total'] or 0,
        'pending_payments': Payment.objects.filter(
            status__in=['pending', 'partial']
        ).aggregate(
            total=Coalesce(
                Sum(F('amount') - F('paid_amount'), output_field=DecimalField()),
                Value(0, output_field=DecimalField())
            )
        )['total'] or 0,
        'paid_payments': Payment.objects.filter(
            status='paid'
        ).aggregate(total=Coalesce(Sum('paid_amount'), Value(0, output_field=DecimalField())))['total'] or 0,
    }
    
    return Response(stats)


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def export_payments_excel(request):
    # Filter payments
    payments_queryset = Payment.objects.all()
    status_param = request.GET.get('status')
    branch = request.GET.get('branch')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    if status_param:
        payments_queryset = payments_queryset.filter(status=status_param)
    if branch:
        payments_queryset = payments_queryset.filter(branch_id=branch)
    if start_date and end_date:
        payments_queryset = payments_queryset.filter(created_at__date__range=[start_date, end_date])
    
    # Filter expenses
    expenses_queryset = Expense.objects.all()
    if branch:
        expenses_queryset = expenses_queryset.filter(branch_id=branch)
    if start_date and end_date:
        expenses_queryset = expenses_queryset.filter(expense_date__range=[start_date, end_date])

    wb = Workbook()
    
    # Payments sheet
    ws_payments = wb.active
    ws_payments.title = 'Thanh toán'
    headers_payments = ['ID', 'ID Khách hàng', 'Khách hàng', 'Dịch vụ', 'Chi nhánh', 'Tổng tiền', 'Đã trả', 'Còn lại', 'Trạng thái', 'Phương thức', 'Ngày tạo']
    ws_payments.append(headers_payments)
    for p in payments_queryset.select_related('customer', 'branch').prefetch_related('services'):
        services_names = ', '.join([service.name for service in p.services.all()])
        ws_payments.append([
            p.id,
            p.customer.id,
            p.customer.full_name,
            services_names,
            p.branch.name,
            int(p.amount),
            int(p.paid_amount),
            int(p.remaining_amount),
            p.get_status_display(),
            p.get_payment_method_display(),
            p.created_at.strftime('%d/%m/%Y'),
        ])
    
    # Expenses sheet
    ws_expenses = wb.create_sheet(title='Chi phí')
    headers_expenses = ['ID', 'Tiêu đề', 'Mô tả', 'Danh mục', 'Số tiền', 'Chi nhánh', 'Ngày chi', 'Ngày tạo']
    ws_expenses.append(headers_expenses)
    for e in expenses_queryset.select_related('branch'):
        ws_expenses.append([
            e.id,
            e.title,
            e.description or '',  # Thêm mô tả, để trống nếu không có
            e.get_category_display(),
            int(e.amount),
            e.branch.name,
            e.expense_date.strftime('%d/%m/%Y'),
            e.created_at.strftime('%d/%m/%Y'),
        ])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=financial_report.xlsx'
    return response


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def export_payments_pdf(request):
    queryset = Payment.objects.all()
    status_param = request.GET.get('status')
    branch = request.GET.get('branch')
    if status_param:
        queryset = queryset.filter(status=status_param)
    if branch:
        queryset = queryset.filter(branch_id=branch)

    buffer = io.BytesIO()
    p = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    y = height - 50
    p.setFont('Helvetica-Bold', 14)
    p.drawString(50, y, 'Danh sách thanh toán')
    y -= 30
    p.setFont('Helvetica-Bold', 10)
    p.drawString(50, y, 'KH | DV | Tổng | Đã trả | Còn lại | TT | CN')
    y -= 15
    p.setFont('Helvetica', 10)
    for r in queryset.select_related('customer', 'branch').prefetch_related('services'):
        services_names = ', '.join([service.name for service in r.services.all()])
        row = f"{r.customer.full_name} | {services_names} | {int(r.amount):,} | {int(r.paid_amount):,} | {int(r.remaining_amount):,} | {r.status} | {r.branch.name}"
        p.drawString(50, y, row)
        y -= 14
        if y < 50:
            p.showPage()
            y = height - 50
    p.showPage()
    p.save()
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=payments.pdf'
    return response


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def export_expenses_excel(request):
    queryset = Expense.objects.all()
    branch = request.GET.get('branch')
    category = request.GET.get('category')
    if branch:
        queryset = queryset.filter(branch_id=branch)
    if category:
        queryset = queryset.filter(category=category)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Expenses'
    headers = ['ID', 'Tiêu đề', 'Danh mục', 'Số tiền', 'Chi nhánh', 'Ngày chi']
    ws.append(headers)
    for e in queryset.select_related('branch'):
        ws.append([
            e.id,
            e.title,
            e.category,
            int(e.amount),
            e.branch.name,
            e.expense_date.strftime('%Y-%m-%d'),
        ])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=expenses.xlsx'
    return response


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def merge_payments(request):
    """Gộp các Payment records của cùng một appointment thành một record duy nhất"""
    try:
        from django.db import transaction
        from appointments.models import Appointment
        
        # Tìm tất cả appointments có nhiều hơn 1 payment
        appointments_with_multiple_payments = []
        
        for appointment in Appointment.objects.all():
            payments = Payment.objects.filter(appointment=appointment)
            if payments.count() > 1:
                appointments_with_multiple_payments.append({
                    'appointment': appointment,
                    'payments': payments,
                    'count': payments.count()
                })
        
        if not appointments_with_multiple_payments:
            return Response({
                'success': True,
                'message': 'Không có appointment nào có nhiều Payment records',
                'merged_count': 0,
                'deleted_count': 0
            })
        
        total_merged = 0
        total_deleted = 0
        
        for item in appointments_with_multiple_payments:
            appointment = item['appointment']
            payments = item['payments']
            
            try:
                with transaction.atomic():
                    # Chọn Payment đầu tiên làm Payment chính
                    main_payment = payments.first()
                    
                    # Gộp tất cả services vào Payment chính
                    all_services = set()
                    total_amount = 0
                    
                    for payment in payments:
                        all_services.update(payment.services.all())
                        total_amount += payment.amount
                    
                    # Cập nhật Payment chính
                    main_payment.services.set(all_services)
                    main_payment.amount = total_amount
                    main_payment.save()
                    
                    # Xóa các Payment khác
                    other_payments = payments.exclude(id=main_payment.id)
                    deleted_count = other_payments.count()
                    other_payments.delete()
                    
                    total_merged += 1
                    total_deleted += deleted_count
                    
            except Exception as e:
                print(f"Lỗi khi gộp Payment cho appointment {appointment.id}: {e}")
                continue
        
        return Response({
            'success': True,
            'message': f'Đã gộp {total_merged} appointments và xóa {total_deleted} Payment records trùng lặp',
            'merged_count': total_merged,
            'deleted_count': total_deleted
        })
        
    except Exception as e:
        return Response({
            'success': False,
            'error': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def export_expenses_pdf(request):
    queryset = Expense.objects.all()
    branch = request.GET.get('branch')
    category = request.GET.get('category')
    if branch:
        queryset = queryset.filter(branch_id=branch)
    if category:
        queryset = queryset.filter(category=category)

    buffer = io.BytesIO()
    p = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    y = height - 50
    p.setFont('Helvetica-Bold', 14)
    p.drawString(50, y, 'Danh sách chi phí')
    y -= 30
    p.setFont('Helvetica-Bold', 10)
    p.drawString(50, y, 'Tiêu đề | Danh mục | Số tiền | Chi nhánh | Ngày')
    y -= 15
    p.setFont('Helvetica', 10)
    for e in queryset.select_related('branch'):
        row = f"{e.title} | {e.category} | {int(e.amount):,} | {e.branch.name} | {e.expense_date.strftime('%Y-%m-%d')}"
        p.drawString(50, y, row)
        y -= 14
        if y < 50:
            p.showPage()
            y = height - 50
    p.showPage()
    p.save()
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=expenses.pdf'
    return response
